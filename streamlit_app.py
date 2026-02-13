import streamlit as st
import pandas as pd
import re
import io
import os
from datetime import datetime
import fitz
import openpyxl
from openpyxl.utils import get_column_letter
import pytesseract
from PIL import Image

# ===========================
# UTILIDADES GENERALES
# ===========================

def safe_filename(name):
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", name)

def now_stamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def autosize_columns(ws):
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(10, max_len + 2), 60)

def ocr_image_to_text(file_bytes):
    try:
        img = Image.open(io.BytesIO(file_bytes)).convert("RGB")
        return pytesseract.image_to_string(img)
    except:
        return ""

# ===========================
# MÓDULO 1 – PDF → EXCEL (TABLAS)
# ===========================

def transformar_archivos_a_excel(uploaded_files):
    regex_documento = re.compile(r"^(CC|TI|CE|RC|NIT)\s+(\d{5,})\s+(.+)$")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos PDF"
    ws.append(["TipoDoc","NumDoc","Nombre","Col1","Col2","Col3","Col4","Col5","Archivo"])

    filas = archivos = 0

    for uf in uploaded_files:
        archivos += 1
        doc = fitz.open(stream=uf.getvalue(), filetype="pdf")
        tipo = num = nombre = ""

        for page in doc:
            for tabla in page.find_tables():
                for fila in tabla.extract():
                    if not any(fila):
                        continue
                    texto = " ".join(str(c) for c in fila if c)
                    m = regex_documento.match(texto)
                    if m:
                        tipo, num, nombre = m.groups()
                        continue

                    fila_limpia = []
                    for c in fila:
                        if isinstance(c, str):
                            c = c.replace("$","").replace(",","").strip()
                            try: c = float(c)
                            except: pass
                        fila_limpia.append(c)

                    ws.append([tipo,num,nombre] + fila_limpia + [uf.name])
                    filas += 1

        doc.close()

    autosize_columns(ws)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, archivos, filas

# ===========================
# MÓDULO 2 – FIRMAR PDFs
# ===========================

def firmar_pdfs_en_zip(pdfs, firma):
    import zipfile

    firma_bytes = firma.getvalue()
    z = io.BytesIO()

    with zipfile.ZipFile(z, "w", zipfile.ZIP_DEFLATED) as zipf:
        for pdf in pdfs:
            doc = fitz.open(stream=pdf.getvalue(), filetype="pdf")
            page = doc[-1]

            instancias = page.search_for("Firma Prestador")

            # Tamaño de la firma (ajustable)
            firma_width = 140
            firma_height = 55

            if instancias:
                rect_texto = instancias[0]

                # 👉 Colocar la firma SOBRE la línea
                x = rect_texto.x0
                y = rect_texto.y1 - firma_height + 8

                rect_firma = fitz.Rect(
                    x,
                    y,
                    x + firma_width,
                    y + firma_height
                )
            else:
                rect_firma = fitz.Rect(70, 130, 210, 185)

            page.insert_image(rect_firma, stream=firma_bytes)

            buf = io.BytesIO()
            doc.save(buf)
            doc.close()

            zipf.writestr(pdf.name, buf.getvalue())

    z.seek(0)
    return z



# ===========================
# MÓDULO 3 – CANCELADAS
# ===========================

def reprogramar_canceladas_excel(file_bytes):
    import io
    import re
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook

    # 1) Leer archivo (maneja .xls/.xlsx)
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), header=None)
    except Exception:
        df = pd.read_excel(io.BytesIO(file_bytes), header=None, engine="xlrd")

    # 2) Tomar "Impresión" desde B1 (fila 0, col 1)
    impresion_origen = ""
    try:
        if isinstance(df.iloc[0, 1], str):
            impresion_origen = df.iloc[0, 1].strip()
    except Exception:
        impresion_origen = ""

    # ---------- helpers ----------
    def to_dt_mixed(series):
        """
        Convierte una columna a datetime soportando:
        - strings dd/mm/yy, dd/mm/yyyy
        - datetime ya existentes
        - números tipo Excel serial (a veces vienen en .xls)
        """
        s = series.copy()

        # Primero intento normal
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")

        # Si hay muchos NaT y hay números, intento convertir serial Excel
        # (Excel: origen 1899-12-30)
        if dt.notna().sum() == 0:
            num = pd.to_numeric(s, errors="coerce")
            if num.notna().sum() > 0:
                dt2 = pd.to_datetime(num, unit="d", origin="1899-12-30", errors="coerce")
                # me quedo con lo que sirva
                dt = dt2

        return dt

    def best_date_col(candidates):
        best_col = None
        best_count = -1
        best_dt = None
        for c in candidates:
            if c < 0 or c >= df.shape[1]:
                continue
            dt = to_dt_mixed(df.iloc[:, c])
            cnt = dt.notna().sum()
            if cnt > best_count:
                best_count = cnt
                best_col = c
                best_dt = dt
        return best_col, best_dt

    # 3) Detectar columna de Cita y Nueva por “mayor cantidad de fechas válidas”
    # (en distintos reportes se corren)
    cita_col, cita_dt = best_date_col([2, 1, 3, 0])      # C, B, D, A
    nueva_col, nueva_dt = best_date_col([8, 7, 9, 6])    # I, H, J, G

    # Si por alguna razón no detecta Cita, no hay nada que filtrar
    if cita_col is None or cita_dt is None or cita_dt.notna().sum() == 0:
        df_out = pd.DataFrame(columns=["Conse", "Cita", "Nombre", "Telefono", "Nueva", "Doctor", "Anotaciones"])
        # export vacío pero con impresión
        temp_output = io.BytesIO()
        df_out.to_excel(temp_output, index=False, startrow=1)
        temp_output.seek(0)
        wb = load_workbook(temp_output)
        ws = wb.active
        if impresion_origen:
            ws["A1"] = impresion_origen
            ws["A1"].font = openpyxl.styles.Font(bold=True)
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        return final_output, df_out

    # 4) Detectar doctor por bloques (probamos columna B y A, nos quedamos con la que funcione)
    def detect_doctor_col(col_idx):
        doctor_actual = ""
        doctor_series = []
        for _, row in df.iterrows():
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                texto = val.strip()
                if texto.isupper() and "CITAS" not in texto and len(texto) > 5:
                    doctor_actual = texto
            doctor_series.append(doctor_actual)
        return doctor_series

    # elegimos entre col 1 y 0 (B o A) según cuál produce más doctores no vacíos
    doc_b = detect_doctor_col(1) if df.shape[1] > 1 else [""] * len(df)
    doc_a = detect_doctor_col(0)
    doctor_list = doc_b if sum(1 for x in doc_b if x) >= sum(1 for x in doc_a if x) else doc_a

    # 5) Detectar columnas de Nombre y Teléfono (heurística simple)
    # Nombre suele estar cerca (E/F). Teléfono cerca (F/G).
    # Si no, caemos a los índices clásicos.
    nombre_candidates = [5, 4, 6, 3]
    tel_candidates = [6, 5, 4, 7]

    def best_text_col(candidates):
        best = None
        best_cnt = -1
        for c in candidates:
            if c < 0 or c >= df.shape[1]:
                continue
            s = df.iloc[:, c].astype(str)
            # cuenta textos que parezcan nombre (no nan, no header, largo>3)
            cnt = ((s.str.lower() != "nan") & (s.str.len() > 3) & (~s.str.contains("fecha|hora|ident|paciente|telefono|actividad|nueva", case=False, na=False))).sum()
            if cnt > best_cnt:
                best_cnt = cnt
                best = c
        return best

    def best_phone_col(candidates):
        best = None
        best_cnt = -1
        for c in candidates:
            if c < 0 or c >= df.shape[1]:
                continue
            s = df.iloc[:, c].astype(str)
            # cuenta celdas con dígitos (teléfono)
            cnt = s.str.contains(r"\d{6,}", regex=True, na=False).sum()
            if cnt > best_cnt:
                best_cnt = cnt
                best = c
        return best

    nombre_col = best_text_col(nombre_candidates)
    tel_col = best_phone_col(tel_candidates)

    if nombre_col is None:
        nombre_col = 5 if df.shape[1] > 5 else df.shape[1] - 1
    if tel_col is None:
        tel_col = 6 if df.shape[1] > 6 else df.shape[1] - 1

    # 6) Aplicar regla: Nueva en blanco O Nueva <= Cita
    # Si no pudimos detectar Nueva, la tratamos como vacía (incluye)
    if nueva_dt is None:
        nueva_dt = pd.Series([pd.NaT] * len(df))

    mask = cita_dt.notna() & (nueva_dt.isna() | (nueva_dt <= cita_dt))

    # 7) Construir salida
    cita_txt = df.iloc[:, cita_col].astype(str).str.replace("*", "", regex=False).str.strip()
    nueva_txt = df.iloc[:, nueva_col].astype(str).str.strip() if nueva_col is not None else ""

    nombre_txt = df.iloc[:, nombre_col].astype(str).str.strip()
    tel_txt = df.iloc[:, tel_col].astype(str).str.strip()

    anotaciones_txt = ""
    if df.shape[1] > 12:
        anotaciones_txt = df.iloc[:, 12].astype(str).str.strip()

    out_rows = []
    for idx in df.index[mask]:
        nom = nombre_txt.loc[idx]
        if str(nom).lower() == "nan":
            continue
        out_rows.append([
            cita_txt.loc[idx],
            nom,
            tel_txt.loc[idx],
            (nueva_txt.loc[idx] if nueva_col is not None else ""),
            doctor_list[idx],
            (anotaciones_txt.loc[idx] if isinstance(anotaciones_txt, pd.Series) else "")
        ])

    df_out = pd.DataFrame(out_rows, columns=["Cita", "Nombre", "Telefono", "Nueva", "Doctor", "Anotaciones"])
    df_out.insert(0, "Conse", range(1, len(df_out) + 1))

    # 8) Exportar a EXCEL con encabezado "Impresión" en A1
    temp_output = io.BytesIO()
    df_out.to_excel(temp_output, index=False, startrow=1)
    temp_output.seek(0)

    wb = load_workbook(temp_output)
    ws = wb.active
    if impresion_origen:
        ws["A1"] = impresion_origen
        ws["A1"].font = openpyxl.styles.Font(bold=True)

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output, df_out




# ===========================
# MÓDULO 4 – INASISTIDAS
# ===========================

def reprogramar_inasistidas_xls(file_bytes):
    import io
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook

    # -------------------------------------------------
    # 1. Leer archivo origen (para lógica)
    # -------------------------------------------------
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, engine="xlrd")

    # -------------------------------------------------
    # 2. Tomar la fecha/leyenda desde A1 del origen
    # -------------------------------------------------
    encabezado_origen = ""
    try:
        if isinstance(df_raw.iloc[0, 0], str):
            encabezado_origen = df_raw.iloc[0, 0].strip()
    except:
        encabezado_origen = ""

    # -------------------------------------------------
    # 3. LÓGICA QUE YA FUNCIONA (NO TOCADA)
    # -------------------------------------------------
    df = df_raw.copy()

    df["Doctor"] = None
    doctor_actual = None

    for i, row in df.iterrows():
        texto = str(row[0]).strip()
        if texto.isupper() and len(texto.split()) > 1:
            doctor_actual = texto
        df.at[i, "Doctor"] = doctor_actual

    df = df[df[3].notnull() & df[0].notnull()]

    df = df.rename(columns={
        0: "Cita_inici",
        2: "Identifica",
        3: "Nombre_paciente",
        4: "Telefono",
        6: "Nueva_cit"
    })

    df["Cita_inici"] = pd.to_datetime(df["Cita_inici"], errors="coerce")
    df["Nueva_cit"] = pd.to_datetime(df["Nueva_cit"], errors="coerce")

    # ✅ incluir si Nueva_cit está en blanco o <= Cita_inici
    df_filtrado = df[df["Nueva_cit"].isna() | (df["Nueva_cit"] <= df["Cita_inici"])].copy()
    df_filtrado = df_filtrado[df_filtrado["Cita_inici"].notnull()]

    df_filtrado = df_filtrado.reset_index(drop=True)
    df_filtrado.insert(0, "Conse", df_filtrado.index + 1)
    df_filtrado["Anotaciones"] = ""

    # -------------------------------------------------
    # 4. Exportar Excel con encabezado en la fila 1
    # -------------------------------------------------
    temp_out = io.BytesIO()
    df_filtrado.to_excel(temp_out, index=False, startrow=1)
    temp_out.seek(0)

    wb = load_workbook(temp_out)
    ws = wb.active

    if encabezado_origen:
        ws["A1"] = encabezado_origen
        ws["A1"].font = openpyxl.styles.Font(bold=True)

    final_out = io.BytesIO()
    wb.save(final_out)
    final_out.seek(0)

    return final_out, df_filtrado





# ===========================
# UI STREAMLIT
# ===========================

st.set_page_config("Denti Manager Web", layout="centered")
st.title("Denti Manager")

tab1, tab2, tab3, tab4 = st.tabs([
    "📄 PDF → Excel",
    "✍️ Firmar PDFs",
    "🚷 Citas Canceladas",
    "🔄 Citas Inasistidas"
])

with tab1:
    files = st.file_uploader("PDFs", type=["pdf"], accept_multiple_files=True)
    if st.button("Procesar PDFs", key="btn_pdf", disabled=not files):
        out, a, f = transformar_archivos_a_excel(files)
        st.success(f"Archivos: {a} | Filas: {f}")
        st.download_button("Descargar Excel", out, f"PDF_{now_stamp()}.xlsx", key="dl_pdf")

with tab2:
    firma = st.file_uploader("Firma", type=["png","jpg"], key="firma")
    pdfs = st.file_uploader("PDFs", type=["pdf"], accept_multiple_files=True, key="pdfs")
    if st.button("Firmar PDFs", key="btn_firmar", disabled=not (firma and pdfs)):
        z = firmar_pdfs_en_zip(pdfs, firma)
        st.download_button("Descargar ZIP", z, f"FIRMADOS_{now_stamp()}.zip", key="dl_zip")

with tab3:
    f = st.file_uploader("Canceladas", type=["xls","xlsx"], key="cancel")
    if st.button("Generar Canceladas", key="btn_cancel", disabled=not f):
        out, df = reprogramar_canceladas_excel(f.getvalue())
        st.dataframe(df.head())
        st.download_button("Descargar", out, f"CANCELADAS_{now_stamp()}.xlsx", key="dl_cancel")

with tab4:
    f = st.file_uploader("Inasistidas", type=["xls"], key="inasis")
    if st.button("Generar Inasistidas", key="btn_inas", disabled=not f):
        out, df = reprogramar_inasistidas_xls(f.getvalue())
        st.dataframe(df.head())
        st.download_button("Descargar", out, f"INASISTIDAS_{now_stamp()}.xlsx", key="dl_inas")



























